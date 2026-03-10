"""
Data exporters: CSV, JSON, and Excel (XLSX).

All exporters share the same interface::

    exporter = CSVExporter(output_dir)
    path = exporter.export(restaurants, filename_stem="grabfood_jakarta_20260309")
    print(f"Exported to {path}")
"""

from __future__ import annotations

import csv
import json
from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path
from typing import List

from scraper.models import Restaurant
from scraper.utils.logger import get_logger

logger = get_logger(__name__)


class BaseExporter(ABC):
    """Abstract base for all export formats."""

    def __init__(self, output_dir: Path) -> None:
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    @abstractmethod
    def export(self, restaurants: List[Restaurant], filename_stem: str) -> Path:
        """Write restaurants to a file and return the path."""

    def _default_stem(self, platform: str = "") -> str:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{platform}_{ts}" if platform else ts


class CSVExporter(BaseExporter):
    """Export restaurants to a flat CSV file."""

    FIELDNAMES = [
        "platform", "restaurant_id", "name", "rating", "review_count",
        "delivery_time", "delivery_fee", "minimum_order", "cuisines",
        "price_range", "city", "district", "address", "latitude", "longitude",
        "is_open", "is_promoted", "promo_label", "menu_item_count", "url",
        "scraped_at",
    ]

    def export(self, restaurants: List[Restaurant], filename_stem: str = "") -> Path:
        stem = filename_stem or self._default_stem()
        path = self.output_dir / f"{stem}.csv"

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=self.FIELDNAMES, extrasaction="ignore")
            writer.writeheader()
            for r in restaurants:
                writer.writerow(r.to_flat_dict())

        logger.info("CSV exported: %s (%d rows)", path, len(restaurants))
        return path


class JSONExporter(BaseExporter):
    """Export restaurants to a JSON file (array of objects)."""

    def export(self, restaurants: List[Restaurant], filename_stem: str = "") -> Path:
        stem = filename_stem or self._default_stem()
        path = self.output_dir / f"{stem}.json"

        data = [r.model_dump(mode="json") for r in restaurants]
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)

        logger.info("JSON exported: %s (%d records)", path, len(restaurants))
        return path


class ExcelExporter(BaseExporter):
    """Export restaurants to an XLSX file (requires openpyxl)."""

    def export(self, restaurants: List[Restaurant], filename_stem: str = "") -> Path:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for Excel export: pip install openpyxl"
            ) from exc

        stem = filename_stem or self._default_stem()
        path = self.output_dir / f"{stem}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Restaurants"

        headers = CSVExporter.FIELDNAMES
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, r in enumerate(restaurants, start=2):
            flat = r.to_flat_dict()
            for col, key in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col, value=flat.get(key))

        # Auto-size columns
        for col in ws.columns:
            max_length = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        wb.save(path)
        logger.info("Excel exported: %s (%d rows)", path, len(restaurants))
        return path
